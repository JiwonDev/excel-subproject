import os
import threading
from datetime import datetime

import numpy
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from win32api import GetShortPathName


class ExcelConvertor(object) :
    def header_func_half_empty(li) :
        return False if (
                                len(li) /
                                2) < len(
            list(
                filter(
                    lambda x :(
                                      x is None) or str(x) == "" or str(x).isspace(),
                    li))) else True

    def header_func_first_line(li) :
        return True if any(li) else False

    @staticmethod
    def execute_file(dir: str) :
        if dir :
            if os.path.isfile(dir) :
                sdir = GetShortPathName(dir)
                threading._start_new_thread(os.system, (sdir,))

    @staticmethod
    def workbook_to_dataframe(
            workbook: Workbook,
            sheet: [
                str,
                int,
                Worksheet] = None,
            header_func=header_func_first_line,
            remove_empty_column=True,
            errorMsg: str = datetime.today().strftime("%Y/%m/%d %H:%M:%S")) -> DataFrame :
        """

        :param workbook: 저장할 엑셀 workbook [Openpyxl.Workbook]
        :param sheet: 인덱스값, [str, int, openpyxl.Workbook]
        :param header_func: 머리글 조건
        :param errorMsg: 에러가 일어났을 때 예외에 추가할 메시지
        :return: pandas.Dataframe
        """

        currentSheet: Worksheet

        if sheet is None :
            # 이름을 입력하지 않았다면 현재 활성화 중인 창
            currentSheet = workbook.active
        else :
            if isinstance(sheet, Worksheet) :
                currentSheet = sheet

            elif isinstance(sheet, str) :
                # (str) 입력 시 해당 Sheet명 탐색
                if sheet in workbook.sheetnames :
                    currentSheet = workbook[sheet]
                else :
                    raise IndexError(
                        f'[{sheet}]은 존재하지 않는 시트 이름입니다.\n>> {errorMsg}')

            elif isinstance(sheet, int) :
                # (int) 입력시 해당 Index 탐색
                if 0 <= sheet < len(workbook.sheetnames) :
                    currentSheet = workbook[workbook.sheetnames[sheet]]
                else :
                    raise IndexError(
                        f'[{sheet}]은 존재하지 않는 시트 인덱스입니다.\n>> {errorMsg}')
            else :
                return None

        data = currentSheet.values
        column = []
        for row in data :
            if header_func(row) :
                column = row
                break
        else :  # for-else
            raise IndexError(f"해당 엑셀 파일에 올바른 머리글이 존재하지 않습니다.\n>> {errorMsg}")
        df = DataFrame(data=data, columns=column)
        if remove_empty_column :
            empty_col = []
            for col in list(df) :
                if str(col).isspace() or col == numpy.NaN :
                    empty_col.append(col)
                df.drop(empty_col, axis=1, inplace=True)
        return df

    @staticmethod
    def dataframe_to_workbook(
            df: DataFrame,
            target: Workbook = None,
            sheetTitle: str = "",
            sheetIndex: int = -1,
            add_index=False,
            add_header=True,
            digit_to_integer=True) -> Workbook :
        workbook = None

        # Workbook 생성, 기존의 workbook이 있다면 거기에 추가.
        if target is None :
            workbook = Workbook()
            workbook.remove(workbook.active)  # 기본 생성 sheet 삭제
        else :
            workbook = target

        # Sheet 생성
        title, index = None, None
        if sheetTitle :
            title = sheetTitle
        if sheetIndex >= 0 :
            index = sheetIndex
        sheet = workbook.create_sheet(title, index)

        from openpyxl.utils.dataframe import dataframe_to_rows
        for row in dataframe_to_rows(df, index=add_index, header=add_header) :
            if any(row) :
                if digit_to_integer :
                    for i, x in enumerate(row, start=0) :
                        if str(x).isdigit() :
                            row[i] = x
                sheet.append(row)

        return workbook

    @staticmethod
    def excel_to_workbook(fileDir) -> Workbook :
        """
        [.xls] 인 경우 xlrd로 읽어들여 openpyxl.Workbook 으로 변환.
        [.xlsx] 인 경우 openpyxl로 읽어들임.
        :param fileDir: str, 파일명
        :return: openpyxl.Workbook
        """

        import warnings
        warnings.simplefilter("ignore")
        # UserWarning: Workbook contains no default style, apply openpyxl's default
        # 엑셀파일에 기본 스타일이 없는경우 생기는 오류를 무시하기 위해서 잠깐 꺼둠.
        # return 하기 전에 다시 복구함. 임시로 끄는 것

        book_xlsx = None
        if "." not in fileDir :
            fileDir += ".xlsx"  # 확장자를 입력하지 않았다면

        if fileDir.lower().endswith(".xlsx") :
            book_xlsx = load_workbook(fileDir)

        elif fileDir.lower().endswith(".xls") :
            from xlrd import open_workbook

            book_xlsx = Workbook()
            book_xls = open_workbook(fileDir)
            sheet_names = book_xls.sheet_names()

            for sheet_index in range(0, len(sheet_names)) :
                sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
                if sheet_index == 0 :
                    sheet_xlsx = book_xlsx.active
                    sheet_xlsx.title = sheet_names[sheet_index]
                else :
                    sheet_xlsx = book_xlsx.create_sheet(
                        title=sheet_names[sheet_index])

                for row in range(0, sheet_xls.nrows) :
                    for col in range(0, sheet_xls.ncols) :
                        sheet_xlsx.cell(row=row + 1, column=col + 1).value = \
                            sheet_xls.cell_value(row, col)
        else :
            raise TypeError(f"{fileDir}은 엑셀 파일 형식(.xlsx, xls)이 아닙니다.")
        book_xlsx.close()  # 객체는 유지됨. 파일이 열려있다면 닫음.

        warnings.simplefilter("default")
        # warn 다시 활성화.

        return book_xlsx
