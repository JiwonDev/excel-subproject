from collections import defaultdict

import pandas as pd

from Excel.ExcelConvertor import ExcelConvertor
from Excel.FileExcel import FileExcel


def main() :
    result_dir = ""  # 결과를 저장할 경로
    result_name = ""  # 결과 파일명
    input_dir = ''  # E.xlsx 파일 경로/이름

    print("입력 파일을 읽어들이는 중...\n")
    e = FileExcel(input_dir)

    print("읽어들인 파일을 데이터프레임으로 변환 중...\n")
    word_dict = defaultdict(list)
    key_names = {}
    firstline = ['영어표제항']
    header = ['표제어']
    for index, sheetname in enumerate(e.sheets.keys()) :
        df = e.getDataframe(sheetname)
        print(f"'{sheetname}' 시트를 합치는 중 [{index + 1}/{e.sheetCount}]...\n")
        firstline.append(f"{sheetname}")
        firstline.append("")
        header.append(f"한글풀이")
        header.append(f"활용 및 숙어")

        for i in range(0, len(df)) :
            t = list(df.loc[i])
            key = "".join(str(t[0]).split()).lower()  # 모든 공백문자 제거

            if not word_dict[key] :  # 처음 보는 표제어라면
                word_dict[key] = [" " for i in range(e.sheetCount * 2)]
                key_names[key] = t[0]  # 표제어 원문(공백 제거 하기전)

            word_dict[key][index * 2] = t[1]  # 한글풀이(2열)
            word_dict[key][index * 2 + 1] = ' '.join(item for item in t[2 :] if item)  # 그 외(3열)

    print(f"결과 파일({result_name})을 저장하는 중...\n")
    data = [[key_names[key]] + line for key, line in word_dict.items()]
    df = pd.DataFrame(data, columns=header)
    workbook = ExcelConvertor.dataframe_to_workbook(df, sheetTitle="결과물")

    # 엑셀 스타일 변경
    sh = workbook.active

    # 첫줄 생성
    sh.insert_rows(0)
    print("엑셀 스타일을 변경하는 중...")
    for col, val in enumerate(firstline, start=1) :
        sh.cell(row=1, column=col).value = val

        from openpyxl.utils import get_column_letter
        sh.column_dimensions[get_column_letter(col)].width = 15  # 기본 셀 너비 지정

        if (col - 1) % 2 == 0 and col > 1 :  # 셀 병합
            sh.merge_cells(start_row=1, end_row=1, start_column=col - 1, end_column=col)

    # # 폰트 및 정렬
    # for r in range(1, sh.max_row + 1):
    #     for c in range(1, sh.max_column + 1):
    #         cell = sh.cell(row=r, column=c)
    #         if r == 1 or r == 2:
    #             cell.font = Font(name="나눔바른고딕 옛한글",size=12)
    #             cell.alignment = Alignment(horizontal='center')
    #         else:
    #             cell.font = Font(name="나눔바른고딕 옛한글")

    workbook.save(result_dir + result_name)
    print("완료")


if __name__ == "__main__" :
    main()
