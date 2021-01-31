from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

import Excel.FileBasic
from Excel.ExcelConvertor import ExcelConvertor


class FileExcel(Excel.FileBasic.FileBasic) :
    def __init__(self, path, readExcelData=True) :
        super(FileExcel, self).__init__(path)
        self._workbook = None
        self._sheets = None
        self._sheetCount = 0
        self._isReadExcelData = False

        if readExcelData :
            self.readExcelData()

    @property
    def workbook(self) -> Workbook :
        if self._isReadExcelData :
            return self._workbook
        else :
            raise ValueError("읽어들인 엑셀 파일이 없습니다. readExcelData()를 실행 해 주세요.")

    @property
    def sheets(self) -> dict :
        if self._isReadExcelData :
            return self._sheets
        else :
            raise ValueError("읽어들인 엑셀 파일이 없습니다. readExcelData()를 실행 해 주세요.")

    @property
    def sheetCount(self) -> int :
        return self._sheetCount

    def getSheet(self, index: [int, str]) :
        if isinstance(index, str) :
            if index in self.workbook.sheetnames :
                return self.workbook[index]
            else :
                raise IndexError(f'{self.fileName} - [{index}] sheet 은 존재하지 않는 이름입니다.')

        elif isinstance(index, int) :
            if index >= 0 and index < len(self.workbook.sheetnames) :
                return self.workbook[self.workbook.sheetnames[index]]
            else :
                raise IndexError(f'{self.fileName} - [{index}] sheet 은 존재하지 않는 인덱스입니다.')

    @property
    def activeSheet(self) -> Worksheet :
        if self._isReadExcelData :
            return self.workbook.active
        else :
            raise ValueError("읽어들인 엑셀 파일이 없습니다. readExcelData()를 실행 해 주세요.")

    def readExcelData(self) :
        if self._ext not in [".xlsx", ".xls", '.xlsm'] :
            raise TypeError(
                f"해당 파일은 올바른 엑셀 파일 형식이 아닙니다. \n>> {self._fileName}")
        self._workbook = ExcelConvertor.excel_to_workbook(self._absPath)
        self._sheets = {}
        for s in self._workbook.sheetnames :
            self._sheets[s] = self._workbook[s]
        self._sheetCount = len(self._sheets)
        self._isReadExcelData = True
        self._workbook.close()  # 파일이 열려있다면 파일을 닫음.

    def getDataframe(
            self,
            sheet: [
                str,
                int,
                Worksheet] = None,
            header_func=lambda x :True if any(x) else False) :

        if not self._isReadExcelData :
            raise ValueError("읽어들인 엑셀 파일이 없습니다. readExcelData()를 실행 해 주세요.")
        else :
            df: DataFrame = ExcelConvertor.workbook_to_dataframe(
                self.workbook, sheet, header_func, errorMsg=self.path)

        return df

    def getHeader(self,
                  sheet: [
                      str,
                      int,
                      Worksheet] = None,
                  header_func=ExcelConvertor.header_func_half_empty) :

        currentSheet = None

        workbook = self.workbook
        if sheet is None :
            # 이름을 입력하지 않았다면 현재 활성화 중인 창
            currentSheet = workbook.active
        else :
            if isinstance(sheet, Worksheet) :
                currentSheet = sheet

            elif isinstance(sheet, str) :
                # (str) 입력 시 해당 Sheet명 탐색
                if sheet in workbook.sheetnames :
                    currentSheet = workbook[sheet]
                else :
                    raise IndexError(
                        f'[{sheet}]은 존재하지 않는 시트 이름입니다.\n>> {self.path}')

            elif isinstance(sheet, int) :
                # (int) 입력시 해당 Index 탐색
                if 0 <= sheet < len(workbook.sheetnames) :
                    currentSheet = workbook[workbook.sheetnames[sheet]]
                else :
                    raise IndexError(
                        f'[{sheet}]은 존재하지 않는 시트 인덱스입니다.\n>> {self.path}')

            column: list
            for row in currentSheet.values :
                if header_func(row) :
                    column = row
                    break
            else :  # for-else
                raise IndexError(f"해당 엑셀 파일에 올바른 머리글이 존재하지 않습니다.\n>> {self.path}")

        return column

    def getHeaderSet(self,
                     sheet: [
                         str,
                         int,
                         Worksheet] = None,
                     header_func=ExcelConvertor.header_func_half_empty) :
        return set(self.getHeader(sheet, header_func))
